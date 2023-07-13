import sqlite3 as sq
import pandas as pd
import pyodbc
import sqlalchemy
import threading
from datetime import datetime
import sys
from openpyxl import load_workbook
from xlrd import open_workbook
import xlrd

def pipeline():
  
    eng = sqlalchemy.create_engine('mssql+pyodbc://moha:123456asd@LAPTOP-ML1CIVBL\TEW_SQLEXPRESS/temper?driver=ODBC+Driver+17+for+SQL+Server')

    wb = open_workbook('C:/Users/medom/Desktop/Book2.xlsm')
    sheet = wb.sheets()[0]
    keys = [sheet.cell(6, col_index).value for col_index in range(sheet.ncols)]

    dict_list = []
    for row_index in range(7, sheet.nrows):
        d = {keys[col_index]: sheet.cell(row_index, col_index).value 
        for col_index in range(sheet.ncols)}
        dict_list.append(d)
    
    df = pd.DataFrame.from_dict(dict_list)
    df['Time'] =pd.to_datetime(df['Time'], unit='d', origin='1899-12-30').dt.round('S')
    df['Time'] = pd.to_datetime(df['Time']).dt.time


    data_len = len(df["Time"]) #13
    result =  pd.read_sql("select count(Time) as data_len from sensor_data ", con=eng)
    sql_len = result['data_len'][0]

    if  sql_len > 0:
        print(sql_len)
    
    if data_len > sql_len:
        new_rows = data_len - sql_len
        data = df.loc[:new_rows]
    elif data_len == sql_len:
        print("done")
        sys.exit()
    

    df.to_sql(name="sensor_data", con=eng,if_exists='append',index=False)

    print('Entry Done for ',datetime.now())


def runningTime():
    interval = 25  # seconds
    threading.Timer(interval, runningTime).start()
    pipeline()


if __name__ == '__main__':
    runningTime()